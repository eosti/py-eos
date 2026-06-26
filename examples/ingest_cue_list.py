"""Imports an Excel cuesheet into Eos.

Supports flags, scenes, spotlights, etc.
"""

import argparse
import logging
import time
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, DEFAULT_FONT
from openpyxl.utils import get_column_letter

from eos import Cue, EosSLIP

logger = logging.getLogger(__name__)
# Before running this, make a Q1 with hard zeroes and everything set to preset home


@dataclass
class SpotAction:
    """Container for a spotlight action."""

    intensity: int | None
    character: int | None
    notes: str | None


@dataclass
class CuelistCue:
    """Container for a cue."""

    number: int
    label: str
    flags: list[str]
    cue_time: float | None = None
    spot_1: SpotAction | None = None
    spot_2: SpotAction | None = None
    sfx: str | None = None
    fx: str | None = None


def text_file(path: str) -> str:
    """Check if a file is valid or not."""
    if not Path(path).is_file():
        raise argparse.ArgumentTypeError("Path is not a valid file")

    return path


def generate_blank_cuelist(num_spots=0) -> None:
    headers = [
        "Cue",
        "Pg",
        "Placement",
        "Notes",
        "Time"
    ]
    header_widths = [6.5, 3.75, 40, 30, 4.5]
    flag_headers = [
        "Cue",
        "Scene"
    ]
    flag_widths = [6, 25]
    spot_headers = [
        "Intens",
        "Character",
        "Notes"
    ]
    spot_widths = [6, 12, 6]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cuelist"

    col_idx = 1
    for h in headers:
        c = ws.cell(row=1, column=col_idx, value=h)
        if h in ["Cue", "Pg"]:
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
        if h in ["Cue"]:
            c.font = Font(bold=True)
        ws.merge_cells(start_row=1, end_row=2, start_column=col_idx, end_column=col_idx)
        col_idx += 1

    c = ws.cell(row=1, column=col_idx, value="Flags")
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True)
    ws.merge_cells(start_row=1, end_row=1, start_column=col_idx, end_column=col_idx + len(flag_headers) - 1)
    for h in flag_headers:
        ws.cell(row=2, column=col_idx, value=h)
        col_idx += 1

    for spotnum in range(1, num_spots + 1):
        c = ws.cell(row=1, column=col_idx, value=f"Spot {spotnum}")
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True)
        ws.merge_cells(start_row=1, end_row=1, start_column=col_idx, end_column=col_idx + len(spot_headers) - 1)
        for h in spot_headers:
            ws.cell(row=2, column=col_idx, value=h)
            col_idx += 1

    total_cols = len(headers) + len(flag_headers) + len(spot_headers) * num_spots
    for col in ws.iter_cols(min_row=2, min_col=1, max_col=total_cols):
        col[-1].border = Border(bottom=Side(style="thin"))

    all_widths = header_widths + flag_widths + spot_widths * num_spots

    for col in range(1, total_cols):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = all_widths[col - 1]

    ws.freeze_panes = "A3"
    DEFAULT_FONT.name = "Atkinson Hyperlegible Next"
    wb.save("cuelist.xlsx")


def parse_cuelist(excel_file: str) -> list[CuelistCue]:
    all_cues = []

    with Path(excel_file).open(mode="rb") as f:
        cuelist = pd.read_excel(f, header=[0, 1])

    cuelist = cuelist.rename(columns=lambda x: x if "Unnamed" not in str(x) else "")
    cuelist.columns = [" ".join(a).strip() for a in cuelist.columns.to_flat_index()]
    cuelist["Pg"] = pd.to_numeric(cuelist["Pg"], downcast="integer", errors="coerce")

    if "Spot 1 Intens" not in cuelist.columns:
        cuelist["Spot 1 Intens"] = np.nan

    if "Spot 2 Intens" not in cuelist.columns:
        cuelist["Spot 2 Intens"] = np.nan

    if "SFX" not in cuelist.columns:
        cuelist["SFX"] = np.nan

    if "FX" not in cuelist.columns:
        cuelist["FX"] = np.nan

    convert_dict = {
        "Pg": "Int64",
        "Cue": float,
        "Spot 1 Intens": "Int64",
        "Spot 2 Intens": "Int64",
    }
    cuelist = cuelist.astype(convert_dict)

    for _, row in cuelist.iterrows():
        if not pd.isna(row['pg']):
            cue_label = f"Pg. {row['Pg']}: {row['Placement']}"
        else:
            cue_label = f"{row['Placement']}"
        if not pd.isna(row["Notes"]):
            cue_label += f" ({row['Notes']})"

        flags = []
        if not pd.isna(row["Flags Cue"]):
            flags += row["Flags Cue"].split(" ")
        if not pd.isna(row["Flags Scene"]):
            flags.append(f"Sc {row['Flags Scene']}")

        this_cue = CuelistCue(float(row["Cue"]), cue_label, flags)

        if not pd.isna(row["Time"]):
            this_cue.cue_time = row["Time"]

        if not pd.isna(row["SFX"]):
            this_cue.sfx = row["SFX"]

        if not pd.isna(row["FX"]):
            this_cue.fx = row["FX"]

        for spotnum in [1, 2]:
            continue
            if (
                (not pd.isna(row[f"Spot {spotnum} Intens"]))
                or (not pd.isna(row[f"Spot {spotnum} Notes"]))
                or (not pd.isna(row[f"Spot {spotnum} Character"]))
            ):
                if pd.isna(row[f"Spot {spotnum} Intens"]):
                    intensity = None
                else:
                    intensity = row[f"Spot {spotnum} Intens"]

                if pd.isna(row[f"Spot {spotnum} Notes"]):
                    notes = None
                else:
                    intensity = row[f"Spot {spotnum} Notes"]

                if pd.isna(row[f"Spot {spotnum} Character"]):
                    character = None
                else:
                    character = row[f"Spot {spotnum} Character"]

                action = SpotAction(intensity, character, notes)
                setattr(this_cue, f"spot_{spotnum}", action)

        all_cues.append(this_cue)

    logger.info("Parsed %i cues", len(all_cues))
    return all_cues


def get_all_characters(cues: list[CuelistCue]) -> set[str]:
    all_characters = []
    for c in cues:
        for s in ["spot_1", "spot_2"]:
            spot = getattr(c, s)
            if spot is None:
                continue
            all_characters.append(spot.character)

    sorted_characters = sorted(set(all_characters))
    logger.info("Parsed %i characters", len(sorted_characters))
    return sorted_characters


def generate_character_fps(eos: EosSLIP, characters: set[str], start_idx=400) -> None:
    for idx, val in enumerate(characters):
        eos.send_command(f"Focus_Palette {start_idx + idx} Label {val} #")


def write_cue(eos: EosSLIP, cue: CuelistCue) -> None:
    this_cue = Cue(1, cue.number)
    eos.record_cue(this_cue)
    time.sleep(0.05)
    eos.label_cue(this_cue, cue.label)

    if cue.cue_time is not None:
        eos.set_time(this_cue, cue.cue_time)

    if "I" in cue.flags:
        eos.intensity_block_cue(this_cue)
    if "B" in cue.flags:
        eos.block_cue(this_cue)
    if "A" in cue.flags:
        eos.assert_cue(this_cue)
    if "M" in cue.flags:
        mark_part = eos.record_part(this_cue, 20)
        eos.mark_cue(mark_part)
        eos.label_cue(mark_part, "--- MARK ---")
    if "Mh" in cue.flags:
        mark_part = eos.record_part(this_cue, 20)
        eos.mark_high_cue(mark_part)
        eos.label_cue(mark_part, "--- MARK ---")
    if "Ml" in cue.flags:
        mark_part = eos.record_part(this_cue, 20)
        eos.mark_low_cue(mark_part)
        eos.label_cue(mark_part, "--- MARK ---")
    if next((s for s in cue.flags if "Sc" in s), None):
        scene_marker = next(s for s in cue.flags if "Sc" in s).split(" ", 1)[1]
        eos.add_scene(this_cue, scene_marker)
    if cue.sfx is not None:
        sfx_part = eos.record_part(this_cue, 19)
        eos.label_cue(sfx_part, f"SFX: {cue.sfx}")
    if cue.fx is not None:
        fx_part = eos.record_part(this_cue, 18)
        eos.label_cue(fx_part, f"FX: {cue.fx}")


def write_spot_diff_cuelist(
    eos: EosSLIP, cue: CuelistCue, characters: set[str], spot_idx=400
) -> None:
    """Spotidx is the offset used for presets and cuelists.
    lots of assumptions here that your spots are 401 and 402.
    """
    for spotnum in [1, 2]:
        if getattr(cue, f"spot_{spotnum}") is None:
            continue
        spot_cue = Cue(spot_idx + spotnum, cue.number)
        eos.record_cue(spot_cue)
        cue_label = ""
        if getattr(cue, f"spot_{spotnum}").character is not None:
            fp = spot_idx + characters.index(getattr(cue, f"spot_{spotnum}").character)
            eos.send_command(f"40{spotnum} Focus_Palette {fp} #")
            cue_label = getattr(cue, f"spot_{spotnum}").character
        if getattr(cue, f"spot_{spotnum}").intensity is not None:
            eos.send_command(
                f"40{spotnum} @ " + str(getattr(cue, f"spot_{spotnum}").intensity) + " #"
            )

        if cue.cue_time is not None:
            eos.set_time(spot_cue, cue.cue_time)

        if getattr(cue, f"spot_{spotnum}").notes is not None:
            cue_label += " | " + getattr(cue, f"spot_{spotnum}").notes

        eos.label_cue(spot_cue, cue_label)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", help="Excel cue list", type=text_file, nargs="?", default=None)
    parser.add_argument("--generate", help="Generates a cuelist", action="store_true")

    args = parser.parse_args()
    logging.basicConfig(level="INFO")

    if args.generate:
        generate_blank_cuelist()
        exit()

    all_cues = parse_cuelist(args.excel)
    all_characters = get_all_characters(all_cues)

    eos = EosSLIP("localhost", 3032)
    eos.live()
    eos.clear_cmd_line()

    generate_character_fps(eos, all_characters)

    for i in all_cues:
        write_cue(eos, i)
        write_spot_diff_cuelist(eos, i, all_characters)


if __name__ == "__main__":
    main()
